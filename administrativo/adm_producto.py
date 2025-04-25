import io
import json
import os
import sys
import openpyxl
import traceback
import xlsxwriter
from xlwt import easyxf
from datetime import datetime, timedelta

import xlwt
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse, FileResponse
from django.template.loader import get_template

from administrativo.commonviews import adduserdata
from administrativo.funciones import MiPaginador, null_to_decimal, generar_nombre
from administrativo.forms import AddProductoForm, AddDescripcionForm, AddSubCategoriaForm, EspecificacionProductoForm, \
    AddStockForm, ImagenProductoForm, ImagenesProductoForm, ImportarProductoForm
from administrativo.models import Producto, CategoriaProducto, SubCategoriaProducto, TipoEspecificacion, \
    EspecificacionProducto, Stock, ImagenProducto, KardexProducto

import unicodedata
from django.db.models import Q, Sum, F


@login_required(redirect_field_name='ret', login_url='/loginclinica')
# @secure_module
# @last_access
@transaction.atomic()
def view(request):
    data = {}
    hoy = datetime.now()
    adduserdata(request, data)
    persona = data['persona']
    if request.method == 'POST':
        action = request.POST['action']
        # PRODUCTO -----------------------------------------------------------------------------------------------------
        if action == 'add':
            try:
                form = AddProductoForm(request.POST)
                if form.is_valid():
                    if form.cleaned_data['precio'] < 0:
                        return JsonResponse({'result': True, 'mensaje': 'El precio no puede ser menor a cero.'})
                    nombre = normalizar_texto(form.cleaned_data['nombre'])
                    descripcion = form.cleaned_data['descripcion']
                    #tiporubro = form.cleaned_data['tiporubro']
                    #subcategoria = form.cleaned_data['subcategoria']
                    precio = form.cleaned_data['precio']
                    vigente = form.cleaned_data['vigente']
                    producto = Producto(nombre=nombre, descripcion=descripcion,
                                        precio=precio, vigente=vigente)
                    if 'imagenprincipal' in request.FILES:
                        producto.imagenprincipal = request.FILES['imagenprincipal']
                    producto.save(request)
                    #log(u'Adiciono un producto: %s' % producto, request, 'add')
                    return JsonResponse({'result': False})
                else:
                    return JsonResponse(
                        {'result': True, 'mensaje': 'El formulario no ha sido completado correctamente.'})
            except Exception as ex:
                pass

        if action == 'importarproductos':
            try:
                form = ImportarProductoForm(request.POST, request.FILES)
                if form.is_valid():
                    if 'archivo' in request.FILES:
                        archivo = request.FILES['archivo']
                        workbook = openpyxl.load_workbook(archivo, data_only=True)
                        sheet = workbook.worksheets[0]
                        novedades_registradas = []

                        for row in sheet.iter_rows(min_row=2):
                            cols = [cell.value for cell in row]
                            nombre = str(cols[1]).strip()
                            newproducto = Producto(nombre=nombre, descripcion=nombre, precio=0)
                            newproducto.save(request)
                    return JsonResponse({'result': False})
                else:
                    return JsonResponse(
                        {'result': True, 'mensaje': 'El formulario no ha sido completado correctamente.'})
            except Exception as ex:
                pass

        if action == 'edit':
            try:
                form = AddProductoForm(request.POST)
                if form.is_valid():
                    if form.cleaned_data['precio'] < 0:
                        return JsonResponse({'result': True, 'mensaje': 'El precio no puede ser menor a cero.'})
                    producto = Producto.objects.get(id=request.POST['id'])
                    producto.nombre = form.cleaned_data['nombre'].upper()
                    producto.descripcion = form.cleaned_data['descripcion']
                    #producto.tiporubro = form.cleaned_data['tiporubro']
                    #producto.subcategoria = form.cleaned_data['subcategoria']
                    producto.precio = form.cleaned_data['precio']
                    producto.vigente = form.cleaned_data['vigente']
                    producto.save(request)
                    if 'imagenprincipal' in request.FILES:
                        producto.imagenprincipal = request.FILES['imagenprincipal']
                        producto.save(request)
                    #log(u'Editó un producto: %s' % producto, request, 'edit')
                    return JsonResponse({'result': False})
                else:
                    return JsonResponse({'result': True, 'mensaje': 'El formulario no ha sido completado correctamente.'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Ha ocurrido un error al editar el producto."})

        if action == 'del':
            try:
                producto = Producto.objects.get(pk=request.POST['id'])
                if producto.stocks.filter(status=True).exists():
                    return JsonResponse({"error": True, "mensaje": "No se puede eliminar un producto que está en Stock."})

                producto.status = False
                producto.save(request)
                #log(u'Eliminó un producto: %s' % producto, request, "del")
                return JsonResponse({"error": False, "refresh": True, "mensaje": "Se ha eliminado correctamente el registro."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"error": True, "mensaje": "Ha ocurrido un error al eliminar el producto."})

        if action == 'imprimirkardexfechas':
            try:
                fecha_inicio = request.POST.get('fechainicio')
                fecha_fin = request.POST.get('fechafin')
                if fecha_inicio:
                    fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
                    # Ajuste para incluir desde el inicio del día (00:00:00)
                    fecha_inicio_query = fecha_inicio_dt
                else:
                    fecha_inicio_query = None

                if fecha_fin:
                    fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d')
                    # Ajuste para incluir hasta el final del día (23:59:59.999999)
                    fecha_fin_query = fecha_fin_dt + timedelta(days=1) - timedelta(microseconds=1)
                else:
                    fecha_fin_query = None

                filters = Q(status=True)
                if fecha_inicio_query:
                    filters &= Q(fecha_creacion__gte=fecha_inicio_query)
                if fecha_fin_query:
                    filters &= Q(fecha_creacion__lte=fecha_fin_query)

                stocks = Stock.objects.filter(filters).order_by('-producto_id')

                response = HttpResponse(content_type='application/ms-excel')
                response['Content-Disposition'] = 'attachment; filename="kardexproducto.xls"'
                font_style = xlwt.XFStyle()
                font_style.font.bold = True
                fuentecabecera = easyxf(
                    'font: name Calibri, color-index black, bold on; pattern: pattern solid, fore_colour gray25; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                style2 = easyxf('borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre')

                # Estilos con un azul más oscuro y opaco
                style_header_1 = easyxf(
                    'font: name Calibri, bold on, height 240, color-index white; alignment: horiz centre, vert centre; pattern: pattern solid, fore_colour 23; borders: left thin, right thin, top thin')
                style_header_2 = easyxf(
                    'font: name Calibri, height 200, color-index white; alignment: horiz centre, vert centre; pattern: pattern solid, fore_colour 23; borders: left thin, right thin, bottom thin')

                wb = xlwt.Workbook(encoding='utf-8')
                ws = wb.add_sheet('hoja1')

                row_num = 5
                columns = [
                    ('ID', 2000),
                    ('Usuario', 2000),
                    ('Producto stock', 12000),
                    ('Fecha', 12000),
                    ('Movimiento', 8000),
                    ('Cantidad', 4000),
                    ('Costo', 4000),
                    ('Total', 4000),
                    ('Observación', 8000),
                ]

                # Escribir el título en las filas combinadas
                ws.write_merge(0, 1, 0, 8, 'CLÍNICA SANTA ELENA', style_header_1)
                ws.write_merge(2, 3, 0, 8, 'FLUJO DE MOVIMIENTOS DEL PRODUCTO EN STOCK', style_header_2)

                for col_num in range(len(columns)):
                    ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                    ws.col(col_num).width = columns[col_num][1]

                row_num += 1
                for stock in stocks:
                    listado = KardexProducto.objects.filter(status=True, stock=stock).order_by('-id')
                    for lis in listado.order_by('-fecha_creacion'):
                        ws.write(row_num, 0, str(lis.id), style2)
                        ws.write(row_num, 1, str(lis.usuario_creacion), style2)
                        ws.write(row_num, 2, str(lis.stock), style2)
                        ws.write(row_num, 3, str(lis.fecha_creacion), style2)
                        ws.write(row_num, 4, lis.get_movimiento_display(), style2)
                        ws.write(row_num, 5, lis.cantidad, style2)
                        ws.write(row_num, 6, f"{lis.costo:.2f}", style2)
                        ws.write(row_num, 7, f"{lis.total:.2f}", style2)
                        ws.write(row_num, 8, str(lis.observacion or ' '), style2)
                        row_num += 1
                wb.save(response)
                return response
            except Exception as ex:
                return JsonResponse({'result': False, 'error': str(ex)})

        # ESPECIFICACIONES DEL PRODUCTO --------------------------------------------------------------------------------
        if action == 'verespecificaciones':
            try:
                with transaction.atomic():
                    body_unicode = request.POST.get('registros')
                    body = json.loads(body_unicode)
                    producto = Producto.objects.get(id=int(request.POST.get('id')))


                    return JsonResponse({"result": False})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse(
                    {"result": True, "mensaje": "Ha ocurrido un error al agregar las especificaciones del producto."})

        if action == 'editverespecificaciones':
            try:
                with transaction.atomic():
                    body_unicode = request.POST.get('registros')
                    body = json.loads(body_unicode)
                    producto = Producto.objects.get(id=int(request.POST.get('id')))

                    for dato in body['datos']:
                        if dato['action'] != 'read':
                            tipoespecificacion = TipoEspecificacion.objects.get(id=dato['tipoespecificacion_id'])
                            valor = normalizar_texto(dato['especificacion'])
                            orden = dato['orden']
                            if dato['action'] == 'add':
                                especificacionproducto = EspecificacionProducto.objects.filter(producto=producto, valor=valor, especificacion=tipoespecificacion, orden=orden, status=False).first()
                                if especificacionproducto:
                                    especificacionproducto.status = True
                                    especificacionproducto.save()
                                else:
                                    especificacionproducto = EspecificacionProducto(valor=dato['especificacion'], especificacion=tipoespecificacion, producto=producto, orden=orden)
                                    especificacionproducto.save(request)
                            if dato['action'] == 'edit' and dato['especificacion_id'] != '':
                                especificacionproducto = EspecificacionProducto.objects.get(id=dato['especificacion_id'])
                                especificacionproducto.valor = valor
                                especificacionproducto.especificacion = tipoespecificacion
                                especificacionproducto.orden = orden
                                especificacionproducto.save(request)
                            if dato['action'] == 'del' and dato['especificacion_id'] != '':
                                especificacionproducto = EspecificacionProducto.objects.get(id=dato['especificacion_id'])
                                especificacionproducto.status = False
                                especificacionproducto.save(request)
                    return JsonResponse({"result": False})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Ha ocurrido un error al agregar las especificaciones del producto."})

        if action == 'editespecificacion':
            try:
                especificacionproducto = EspecificacionProducto.objects.get(id=request.POST['id'])
                if especificacionproducto.stocks.exists():
                    return JsonResponse({'result': True,'mensaje': f'No se puede eliminar la especificación "{especificacionproducto}" porque está asociada a uno o más stocks.'})
                else:
                    return JsonResponse({'result': False})
            except Exception as ex:
                pass

        # IMAGENES DEL PRODUCTO ----------------------------------------------------------------------------------------
        if action == 'addimagenes':
            try:
                form = ImagenesProductoForm(request.POST, request.FILES)
                if form.is_valid():
                    producto = Producto.objects.get(id=request.POST['id'])
                    for image in request.FILES.getlist('imagen'):
                        imagen = ImagenProducto(imagen=image, producto=producto)
                        imagen.save()
                    #log(u'Adiciono una o más imágenes al Producto: %s' % producto, request, 'add')
                    return JsonResponse({'result': False})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Por favor seleccione una imágen."})

        if action == 'editimagen':
            try:
                form = ImagenProductoForm(request.POST, request.FILES)
                if form.is_valid():
                    imagen = ImagenProducto.objects.get(pk=request.POST['id'])
                    if 'imagen' in request.FILES:
                        imagen.imagen = request.FILES['imagen']
                        imagen.save(request)
                        #log(u'Edito una imágen del producto: %s' % imagen, request, "edit")
                        return JsonResponse({'result': False})
                else:
                    return JsonResponse({"error": True, "mensaje": "Seleccione una imágen válida."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"error": True, "mensaje": "Ha ocurrido un error al eliminar el producto."})

        if action == 'delimagen':
            try:
                imagen = ImagenProducto.objects.get(pk=request.POST['id'])
                imagen.status = False
                imagen.save(request)
                #log(u'Eliminó una imágen del producto: %s' % imagen, request, "del")
                return JsonResponse({"error": False, "refresh": True, "mensaje": "Se ha eliminado correctamente el registro."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"error": True, "mensaje": "Ha ocurrido un error al eliminar el producto."})

        # CATEGORIA ----------------------------------------------------------------------------------------------------
        if action == 'addcategoria':
            try:
                form = AddDescripcionForm(request.POST)
                if form.is_valid():
                    descripcion = normalizar_texto(form.cleaned_data['descripcion'])
                    if CategoriaProducto.objects.filter(descripcion=descripcion, status=True).exists():
                        return JsonResponse({'result': True, 'mensaje': 'Ya existe un categoría con esa descripción.'})
                    categoria = CategoriaProducto(descripcion=descripcion)
                    categoria.save(request)
                    #log(u'Adiciono una categoria: %s' % categoria, request, 'add')
                    return JsonResponse({'result': False})
                else:
                    return JsonResponse({'result': True, 'mensaje': 'El formulario no ha sido completado correctamente.'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Ha ocurrido un error al adicionar la categoría."})

        if action == 'editcategoria':
            try:
                form = AddDescripcionForm(request.POST)
                if form.is_valid():
                    descripcion = normalizar_texto(form.cleaned_data['descripcion'])
                    if CategoriaProducto.objects.filter(descripcion=descripcion, status=True).exists():
                        return JsonResponse({'result': True, 'mensaje': 'Ya existe un categoría con esa descripción.'})
                    categoria = CategoriaProducto.objects.get(id=request.POST['id'])
                    categoria.descripcion = form.cleaned_data['descripcion'].upper()
                    categoria.save(request)
                    #log(u'Editó una categoria: %s' % categoria, request, 'edit')
                    return JsonResponse({'result': False})
                else:
                    return JsonResponse({'result': True, 'mensaje': 'El formulario no ha sido completado correctamente.'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Ha ocurrido un error al editar la categoría."})

        if action == 'delcategoria':
            try:
                categoria = CategoriaProducto.objects.get(pk=request.POST['id'])
                if categoria.subcategoriaproducto_set.filter(status=True).exists():
                    return JsonResponse({"error": True, "mensaje": "No se puede eliminar una categoría que tiene subcategoría."})
                categoria.status = False
                categoria.save(request)
                #log(u'Eliminó una Categoría: %s' % categoria, request, "del")
                return JsonResponse({"error": False, "refresh": True, "mensaje": "Se ha eliminado correctamente el registro."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"error": True, "mensaje": "Ha ocurrido un error al eliminar la categoría."})

        # SUBCATEGORIA -------------------------------------------------------------------------------------------------
        if action == 'addsubcategoria':
            try:
                form = AddSubCategoriaForm(request.POST)
                if form.is_valid():
                    descripcion = normalizar_texto(form.cleaned_data['descripcion'])
                    categoria = form.cleaned_data['categoria']
                    if SubCategoriaProducto.objects.filter(descripcion=descripcion, status=True, categoriaproducto=categoria).exists():
                        return JsonResponse({'result': True, 'mensaje': 'Ya existe una subcategoría con esa descripción y categoría.'})

                    categoriaproducto = categoria
                    subcategoria = SubCategoriaProducto(
                        descripcion=descripcion,
                        categoriaproducto=categoriaproducto
                    )
                    if 'imagen' in request.FILES:
                        subcategoria.imagen = request.FILES['imagen']
                    subcategoria.save(request)
                    #log(u'Adiciono una categoria: %s' % subcategoria, request, 'add')
                    return JsonResponse({'result': False})
                else:
                    return JsonResponse({'result': True, 'mensaje': 'El formulario no ha sido completado correctamente.'})
            except Exception as ex:
                pass

        if action == 'editsubcategoria':
            try:
                form = AddSubCategoriaForm(request.POST)
                if form.is_valid():
                    descripcion = normalizar_texto(form.cleaned_data['descripcion'])
                    categoria = form.cleaned_data['categoria']
                    if SubCategoriaProducto.objects.filter(descripcion=descripcion, status=True, categoriaproducto=categoria).exclude(id=request.POST['id']).exists():
                        return JsonResponse({'result': True, 'mensaje': 'Ya existe una subcategoría con esa descripción y categoría.'})

                    subcategoria = SubCategoriaProducto.objects.get(id=request.POST['id'])
                    subcategoria.descripcion = descripcion
                    subcategoria.categoriaproducto = categoria
                    subcategoria.save(request)
                    if 'imagen' in request.FILES:
                        subcategoria.imagen = request.FILES['imagen']
                        subcategoria.save(request)
                    #log(u'Editó una subcategoria: %s' % subcategoria, request, 'edit')
                    return JsonResponse({'result': False})
                else:
                    return JsonResponse({'result': True, 'mensaje': 'El formulario no ha sido completado correctamente.'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Ha ocurrido un error al editar la subcategoría."})

        if action == 'delsubcategoria':
            try:
                subcategoria = SubCategoriaProducto.objects.get(pk=request.POST['id'])
                subcategoria.status = False
                subcategoria.save(request)
                #log(u'Eliminó una SubCategoría: %s' % subcategoria, request, "del")
                return JsonResponse({"error": False, "refresh": True, "mensaje": "Se ha eliminado correctamente el registro."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"error": True, "mensaje": "Ha ocurrido un error al eliminar la categoría."})

        # TIPO ESPECIFICACION ------------------------------------------------------------------------------------------
        if action == 'addtipoespecificacion':
            try:
                form = AddDescripcionForm(request.POST)
                if form.is_valid():
                    atributo = normalizar_texto(form.cleaned_data['descripcion'])

                    if TipoEspecificacion.objects.filter(atributo=atributo, status=True).exists():
                        return JsonResponse({'result': True, 'mensaje': 'Ya existe un tipo de especificación con esa descripción.'})

                    tipoespecificacion = TipoEspecificacion(atributo=atributo)
                    tipoespecificacion.save(request)
                    #log(u'Adiciono un tipo de especificación: %s' % tipoespecificacion, request, 'add')
                    return JsonResponse({'result': False, 'mensaje': u'Guardado con exito'})
                else:
                    return JsonResponse({'result': True, 'mensaje': 'El formulario no ha sido completado correctamente.'})
            except Exception as ex:
                return JsonResponse({'result': True, 'mensaje': 'Ha ocurrido un error al adicionar un tipo de especificación.'})

        if action == 'edittipoespecificacion':
            try:
                form = AddDescripcionForm(request.POST)
                if form.is_valid():
                    atributo = normalizar_texto(form.cleaned_data['descripcion'])

                    if TipoEspecificacion.objects.filter(atributo=atributo, status=True).exists():
                        return JsonResponse({'result': True, 'mensaje': 'Ya existe este tipo de especificación.'})

                    tipoespecificacion = TipoEspecificacion.objects.get(id=request.POST['id'])
                    tipoespecificacion.atributo = atributo
                    tipoespecificacion.save(request)
                    #log(u'Editó un tipo de especificación: %s' % tipoespecificacion, request, 'edit')
                    return JsonResponse({'result': False, 'mensaje': 'Ha editado un tipo de especificación excitosamente.'})
                else:
                    return JsonResponse({'result': True, 'mensaje': 'El formulario no ha sido completado correctamente.'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse(
                    {"result": True, "mensaje": "Ha ocurrido un error al editar el tipo de especificación."},
                    safe=False)

        if action == 'deltipoespecificacion':
            try:
                tipoespecificacion = TipoEspecificacion.objects.get(pk=request.POST['id'])
                if EspecificacionProducto.objects.filter(especificacion=tipoespecificacion, status=True).exists():
                    return JsonResponse({"error": True, "mensaje": "No se puede eliminar un tipo de especificación que está en uso."})
                else:
                    tipoespecificacion.status = False
                    tipoespecificacion.save(request)
                    #log(u'Eliminó un tipo de especificación: %s' % tipoespecificacion, request, "del")
                    return JsonResponse({"error": False, "refresh": True, "mensaje": "Se ha eliminado correctamente el registro."})

            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"error": True, "mensaje": "Ha ocurrido un error al eliminar la categoría."})

        # STOCK --------------------------------------------------------------------------------------------------------
        if action == 'addstock':
            try:
                post_data = request.POST.copy()
                form = AddStockForm(post_data)
                if form.is_valid():
                    producto = form.cleaned_data['producto']
                    cantidad = form.cleaned_data['cantidad']

                    if cantidad < 1:
                        return JsonResponse({'result': True, 'mensaje': 'Por favor ingrese una cantidad.'})

                    stock = Stock(
                        producto=producto,
                        cantidad=cantidad,
                        cantidad_inicial=cantidad
                    )
                    stock.save()
                    totalcosto = cantidad * producto.precio
                    newkardex = KardexProducto(stock=stock, movimiento=1, cantidad=cantidad, costo=producto.precio, total=totalcosto)
                    newkardex.save(request)

                    # Registrar en el log
                    #log(u'Adicionó un producto al Stock: %s' % stock, request, 'add')

                    # Respuesta exitosa
                    return JsonResponse({'result': False})
                else:
                    return JsonResponse({'result': True, 'mensaje': 'El formulario no ha sido completado correctamente.'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Ha ocurrido un error al adicionar un producto al Stock."})

        if action == 'editstock':
            try:
                post_data = request.POST.copy()

                form = AddStockForm(post_data)
                if form.is_valid():
                    stock = Stock.objects.get(id=request.POST['id'])


                    stocks_existentes = Stock.objects.filter(producto=stock.producto, status=True).exclude(id=stock.id)


                    # Actualizar el stock si no hay duplicados

                    stock.save(request)
                    #log(u'Editó un producto de stock: %s' % stock, request, 'edit')
                    return JsonResponse({'result': False})
                else:
                    return JsonResponse({'result': True, 'mensaje': 'Es necesario que ingrese un stock.'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, "mensaje": "Ha ocurrido un error al editar el stock."})

        if action == 'delstock':
            try:
                stock = Stock.objects.get(pk=request.POST['id'])
                stock.status = False
                stock.save(request)
                #log(u'Eliminó un Stock: %s' % stock, request, "del")
                return JsonResponse({"error": False, "refresh": True, "mensaje": "Se ha eliminado correctamente el registro."})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"error": True, "mensaje": "Ha ocurrido un error al eliminar el stock."})

        if action == 'aumentarstock':
            try:
                stock = Stock.objects.get(id=request.POST['id'])
                cantidad = int(request.POST['cantidad'])
                observacion = request.POST['observacion'].upper()
                stock.cantidad += cantidad
                producto = stock.producto
                costotal = producto.precio * cantidad
                kardex = KardexProducto(
                    stock=stock,
                    movimiento=1,
                    cantidad=cantidad,
                    costo=producto.precio,
                    observacion=observacion,
                    total=costotal
                )
                stock.save(request)
                kardex.save(request)
                return JsonResponse({"result": False})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, 'mensaje': str(ex)})

        if action == 'registrarsalida':
            try:
                stock = Stock.objects.get(id=request.POST['id'])
                cantidad = int(request.POST['cantidad'])
                observacion = request.POST['observacion'].upper()
                if stock.cantidad >= cantidad:
                    stock.cantidad -= cantidad
                    producto = stock.producto
                    costotal = producto.precio * cantidad
                    kardex = KardexProducto(
                        stock=stock,
                        movimiento=2,
                        cantidad=cantidad,
                        costo=producto.precio,
                        observacion=observacion,
                        total=costotal
                    )
                    stock.save(request)
                    kardex.save(request)
                    return JsonResponse({"result": False})
                return JsonResponse({"result": True, 'mensaje': f'Cantidad disponible a disminuir: {stock.cantidad}'})
            except Exception as ex:
                transaction.set_rollback(True)
                return JsonResponse({"result": True, 'mensaje': str(ex)})

        return JsonResponse({'result': True, 'mensaje': 'Ha ocurrido un error al guardar los datos.'})
    else:
        if 'action' in request.GET:
            data['action'] = action = request.GET['action']

            # PRODUCTO -------------------------------------------------------------------------------------------------
            if action == 'add':
                try:
                    form = AddProductoForm()
                    data['form'] = form
                    data['action'] = 'add'
                    data['id'] = request.GET['id']
                    template = get_template("inventario/adm_productotiendavirtual/modal/form_producto.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': str(ex)})

            if action == 'importarproductos':
                try:
                    form = ImportarProductoForm()
                    data['form'] = form
                    data['action'] = 'importarproductos'
                    template = get_template("inventario/adm_productotiendavirtual/formimportar.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': str(ex)})

            if action == 'edit':
                try:
                    data['id'] = request.GET['id']
                    data['action'] = 'edit'
                    producto = Producto.objects.get(pk=request.GET['id'])
                    data['form'] = AddProductoForm(
                        initial={'nombre': producto.nombre, 'descripcion': producto.descripcion,
                                 'precio': producto.precio, 'vigente': producto.vigente})
                    template = get_template("inventario/adm_productotiendavirtual/modal/form_producto.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': str(ex)})

            if action == 'verespecificaciones':
                try:
                    producto = Producto.objects.get(pk=request.GET['id'])
                    especificaciones = producto.especificaciones.filter(status=True)
                    descripcion = f'{producto.nombre} - {producto.precio:.2f}$'
                    data['form'] = EspecificacionProductoForm(initial={'producto': descripcion})
                    data['title'] = f'Especificaciones del producto: {producto.nombre}'
                    data['producto'] = producto
                    data['action'] = request.GET['action']
                    data['id'] = request.GET['id']
                    data['especificaciones'] = ''
                    if especificaciones.exists():
                        data['action'] = 'addespecificaciones'
                        datos = []
                        contador = 0
                        for especificacion in especificaciones:
                            tipoespecificacion = TipoEspecificacion.objects.get(pk=especificacion.especificacion_id)
                            datos.append({
                                "id": contador,
                                "tipoespecificacion_id": str(tipoespecificacion.id),
                                "tipoespecificacion": tipoespecificacion.atributo,
                                "especificacion_id": str(especificacion.id),
                                "especificacion": especificacion.valor,
                                "action": "read",
                                "orden": especificacion.orden
                            })
                            contador += 1
                        data['especificaciones'] = datos
                        return render(request, 'inventario/adm_productotiendavirtual/edit_verespecificaciones.html', data)
                    return render(request, 'inventario/adm_productotiendavirtual/verespecificaciones.html', data)
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': str(ex)})

            if action == 'imagenesproducto':
                try:
                    data['producto'] = producto = Producto.objects.get(id=request.GET['id'])
                    data['title'] = f'Imágenes del producto {producto}'
                    request.session['viewactivo'] = 1
                    listado = ImagenProducto.objects.filter(status=True, producto=producto).order_by('-id')
                    paging = MiPaginador(listado, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                            # p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data["url_vars"] = f'&action={action}&id={producto.id}'
                    data['listado'] = page.object_list
                    return render(request, 'inventario/adm_productotiendavirtual/imagenesproducto.html', data)
                except Exception as ex:
                    pass

            if action == 'addimagenes':
                try:
                    data['form'] = ImagenesProductoForm()
                    data['action'] = action
                    data['id'] = request.GET['id']
                    template = get_template("inventario/adm_productotiendavirtual/modal/modal.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'editimagen':
                try:
                    data['form'] = ImagenProductoForm()
                    data['action'] = action
                    data['id'] = request.GET['id']
                    template = get_template("inventario/adm_productotiendavirtual/modal/modal.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            # CATEGORÍA ------------------------------------------------------------------------------------------------
            if action == 'categoriaproducto':
                try:
                    data['title'] = u'Categorías de los Productos'
                    request.session['viewactivo'] = 2
                    search, url_vars, filtro = request.GET.get('s', ''), f'&action={action}', Q(status=True)
                    if search:
                        data['s'] = search
                        filtro = filtro & Q(descripcion__icontains=search)
                        url_vars += f'&s={search}'
                    listado = CategoriaProducto.objects.filter(filtro).order_by('-id')
                    paging = MiPaginador(listado, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                            # p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data["url_vars"] = url_vars
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['listado'] = page.object_list
                    return render(request, 'inventario/adm_productotiendavirtual/categoriaproducto.html', data)
                except Exception as ex:
                    pass

            if action == 'addcategoria':
                try:
                    form = AddDescripcionForm()
                    data['form'] = form
                    data['action'] = 'addcategoria'
                    data['id'] = request.GET['id']
                    template = get_template("inventario/adm_productotiendavirtual/modal/modal.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': str(ex)})

            if action == 'editcategoria':
                try:
                    data['id'] = request.GET['id']
                    data['action'] = 'editcategoria'
                    categoria = CategoriaProducto.objects.get(pk=request.GET['id'])
                    data['form'] = AddDescripcionForm(initial={'descripcion': categoria.descripcion})
                    template = get_template("inventario/adm_productotiendavirtual/modal/modal.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': str(ex)})
            # SUBCATEGORÍA ---------------------------------------------------------------------------------------------
            if action == 'subcategoriaproducto':
                try:
                    data['title'] = u'Subcategorías de los Productos'
                    request.session['viewactivo'] = 3
                    search, url_vars, filtro = request.GET.get('s', ''), f'&action={action}', Q(status=True)
                    if search:
                        data['s'] = search
                        filtro = filtro & Q(descripcion__icontains=search) | Q(categoriaproducto__descripcion__icontains=search)
                        url_vars += f'&s={search}'
                    listado = SubCategoriaProducto.objects.filter(filtro).order_by('-id')
                    paging = MiPaginador(listado, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                            # p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data["url_vars"] = url_vars
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['listado'] = page.object_list
                    # data['listado'] = SubCategoriaProducto.objects.filter(status=True, categoriaproducto__status=True).order_by('-id')
                    return render(request, 'inventario/adm_productotiendavirtual/subcategoriaproducto.html', data)
                except Exception as ex:
                    pass

            if action == 'addsubcategoria':
                try:
                    form = AddSubCategoriaForm()
                    data['form'] = form
                    data['action'] = 'addsubcategoria'
                    data['id'] = request.GET['id']
                    template = get_template("inventario/adm_productotiendavirtual/modal/modal.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': str(ex)})

            if action == 'editsubcategoria':
                try:
                    data['id'] = request.GET['id']
                    data['action'] = 'editsubcategoria'
                    subcategoria = SubCategoriaProducto.objects.get(pk=request.GET['id'])
                    data['form'] = AddSubCategoriaForm(initial={
                        'descripcion': subcategoria.descripcion,
                        'categoria': subcategoria.categoriaproducto})
                    template = get_template("inventario/adm_productotiendavirtual/modal/modal.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': str(ex)})

            # TIPOESPECIFICACION ---------------------------------------------------------------------------------------
            if action == 'tipoespecificacion':
                try:
                    data['title'] = u'Tipo de Especificación de los Productos'
                    request.session['viewactivo'] = 4
                    search, url_vars, filtro = request.GET.get('s', ''), f'&action={action}', Q(status=True)
                    if search:
                        data['s'] = search
                        filtro = filtro & Q(atributo__icontains=search)
                        url_vars += f'&s={search}'
                    listado = TipoEspecificacion.objects.filter(filtro).order_by('-id')
                    paging = MiPaginador(listado, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                            # p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data["url_vars"] = url_vars
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['listado'] = page.object_list
                    # data['listado'] = TipoEspecificacion.objects.filter(status=True).order_by('-id')
                    return render(request, 'inventario/adm_productotiendavirtual/tipoespecificacion.html', data)
                except Exception as ex:
                    pass

            if action == 'addtipoespecificacion':
                try:
                    form = AddDescripcionForm()
                    data['form'] = form
                    data['action'] = 'addtipoespecificacion'
                    data['id'] = request.GET['id']
                    template = get_template("inventario/adm_productotiendavirtual/modal/modal.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': str(ex)})

            if action == 'edittipoespecificacion':
                try:
                    data['id'] = request.GET['id']
                    data['action'] = 'edittipoespecificacion'
                    tipoespecificacion = TipoEspecificacion.objects.get(pk=request.GET['id'])
                    data['form'] = AddDescripcionForm(initial={'descripcion': tipoespecificacion.atributo})
                    template = get_template("inventario/adm_productotiendavirtual/modal/modal.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': str(ex)})

            # REPORTERIA -----------------------------------------------------------------------------------------------

            if action == 'reporteria':
                try:
                    data['title'] = u'Reportería'
                    request.session['viewactivo'] = 3
                    return render(request, 'inventario/adm_productotiendavirtual/reporteria.html', data)
                except Exception as ex:
                    pass

            # STOCK ----------------------------------------------------------------------------------------------------
            if action == 'stockproducto':
                try:
                    data['title'] = u'Stock de los Productos'
                    request.session['viewactivo'] = 5
                    # stock = Stock.objects.filter(status=True).order_by('-id')
                    search, url_vars, filtro = request.GET.get('s', ''), f'&action={action}', Q(status=True)
                    if search:
                        data['s'] = search
                        filtro = filtro & Q(producto__nombre__icontains=search)
                        url_vars += f'&s={search}'
                    listado = Stock.objects.filter(filtro).order_by('-id')
                    paging = MiPaginador(listado, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                            # p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data["url_vars"] = url_vars
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['listado'] = page.object_list
                    data['totalstock'] = listado.aggregate(total_cantidad=Sum('cantidad'))['total_cantidad']
                    return render(request, 'inventario/adm_productotiendavirtual/productostock.html', data)
                except Exception as ex:
                    pass

            if action == 'addstock':
                try:
                    form = AddStockForm()
                    data['form'] = form
                    data['action'] = 'addstock'
                    template = get_template("inventario/adm_productotiendavirtual/modal/modalstock.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': str(ex)})

            if action == 'editstock':
                try:
                    data['id'] = request.GET['id']
                    data['action'] = 'editstock'
                    stock = Stock.objects.get(pk=request.GET['id'])
                    form = AddStockForm()
                    form.fields['producto'].queryset = Producto.objects.filter(id=stock.producto.id)
                    form.initial = {'producto': stock.producto, 'cantidad': stock.cantidad}
                    data['form'] = form
                    datos = []
                    data['seleccionados'] = datos
                    template = get_template("inventario/adm_productotiendavirtual/modal/tablaespecificaciones.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': str(ex)})

            if action == 'cargarespecificaciones':
                try:
                    datos = []
                    producto = Producto.objects.get(id=request.GET['id'])
                    especificaciones = EspecificacionProducto.objects.filter(status=True, producto=producto)
                    for especificacion in especificaciones:
                        datos.append({
                            'especificacion': f'{especificacion.especificacion}: {especificacion.valor}',
                            'id': f'{especificacion.especificacion.id}-{especificacion.id}'
                        })
                    return JsonResponse({"result": True, "datos": datos})
                except Exception as ex:
                    return JsonResponse({"result": False, 'mensaje': str(ex)})

            # VER STOCK ------------------------------------------------------------------------------------------------
            if action == 'verstock':
                try:
                    data['id'] = id = request.GET['id']
                    producto = Producto.objects.get(id=id)
                    data['title'] = f'Stock de: {producto}'
                    request.session['viewactivo'] = 1
                    search, url_vars, filtro = request.GET.get('s', ''), f'&action={action}&id={id}', Q(status=True) & Q(producto=producto)
                    listado = Stock.objects.filter(filtro).order_by('-id')

                    if search:
                        data['s'] = search
                        filtro = filtro & Q(especificacion_producto__especificacion__atributo__icontains=search) \
                                 | Q(especificacion_producto__valor__icontains=search)
                        url_vars += f'&s={search}'
                        listado = Stock.objects.filter(filtro).order_by('-id')

                    paging = MiPaginador(listado, 10)
                    p = 1
                    try:
                        paginasesion = 1
                        if 'paginador' in request.session:
                            paginasesion = int(request.session['paginador'])
                        if 'page' in request.GET:
                            p = int(request.GET['page'])
                            # p = int(request.GET['page'])
                        else:
                            p = paginasesion
                        try:
                            page = paging.page(p)
                        except:
                            p = 1
                        page = paging.page(p)
                    except:
                        page = paging.page(p)
                    request.session['paginador'] = p
                    data['paging'] = paging
                    data["url_vars"] = url_vars
                    data['rangospaging'] = paging.rangos_paginado(p)
                    data['page'] = page
                    data['listado'] = page.object_list
                    return render(request, 'inventario/adm_productotiendavirtual/verstock.html', data)
                except Exception as ex:
                    pass

            if action == 'verespecificacionesstock':
                try:
                    stock = Stock.objects.get(id=request.GET['id'])
                    request.session['viewactivo'] = 5
                    data['listado'] = stock.especificacion_producto.filter(status=True)
                    data['action'] = 'verespecificacionesstock'
                    template = get_template("inventario/adm_productotiendavirtual/modal/tablaespecificaciones.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'kardex':
                try:
                    stock = Stock.objects.get(id=request.GET['id'])
                    data['listado'] = KardexProducto.objects.filter(status=True, stock=stock).order_by('-id')
                    template = get_template("inventario/adm_productotiendavirtual/modal/form_kardex.html")
                    return JsonResponse({"result": True, 'data': template.render(data)})
                except Exception as ex:
                    pass

            if action == 'imprimirkardex':
                try:
                    stock = Stock.objects.get(id=request.GET['id'])
                    listado = KardexProducto.objects.filter(status=True, stock=stock).order_by('-id')
                    response = HttpResponse(content_type='application/ms-excel')
                    response['Content-Disposition'] = 'attachment; filename="kardexproducto.xls"'
                    font_style = xlwt.XFStyle()
                    font_style.font.bold = True
                    fuentecabecera = easyxf('font: name Calibri, color-index black, bold on; pattern: pattern solid, fore_colour gray25; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    style2 = easyxf('borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre')

                    # Estilos con un azul más oscuro y opaco
                    style_header_1 = easyxf('font: name Calibri, bold on, height 240, color-index white; alignment: horiz centre, vert centre; pattern: pattern solid, fore_colour 23; borders: left thin, right thin, top thin')
                    style_header_2 = easyxf('font: name Calibri, height 200, color-index white; alignment: horiz centre, vert centre; pattern: pattern solid, fore_colour 23; borders: left thin, right thin, bottom thin')

                    wb = xlwt.Workbook(encoding='utf-8')
                    ws = wb.add_sheet('hoja1')
                    row_num = 5
                    columns = [
                        ('ID', 2000),
                        ('Usuario', 2000),
                        ('Producto stock', 12000),
                        ('Fecha', 12000),
                        ('Movimiento', 8000),
                        ('Cantidad', 4000),
                        ('Costo', 4000),
                        ('Total', 4000),
                        ('Observación', 8000),
                    ]

                    # Escribir el título en las filas combinadas
                    ws.write_merge(0, 1, 0, 8, 'CLÍNICA SANTA ELENA', style_header_1)
                    ws.write_merge(2, 3, 0, 8, 'FLUJO DE MOVIMIENTOS DEL PRODUCTO EN STOCK', style_header_2)

                    for col_num in range(len(columns)):
                        ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                        ws.col(col_num).width = columns[col_num][1]

                    row_num += 1
                    for lis in listado.order_by('-fecha_creacion'):
                        ws.write(row_num, 0, str(lis.id), style2)
                        ws.write(row_num, 1, str(lis.usuario_creacion), style2)
                        ws.write(row_num, 2, str(lis.stock), style2)
                        ws.write(row_num, 3, str(lis.fecha_creacion), style2)
                        ws.write(row_num, 4, lis.get_movimiento_display(), style2)
                        ws.write(row_num, 5, lis.cantidad, style2)
                        ws.write(row_num, 6, f"{lis.costo:.2f}", style2)
                        ws.write(row_num, 7, f"{lis.total:.2f}", style2)
                        ws.write(row_num, 8, str(lis.observacion or ' '), style2)
                        row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    return JsonResponse({'result': False, 'error': str(ex)})

            if action == 'imprimirkardexultimasemana':
                try:
                    hoy = datetime.now()
                    ultimo_domingo = hoy - timedelta(days=hoy.weekday() + 1)
                    lunes_pasado = ultimo_domingo - timedelta(days=6)
                    response = HttpResponse(content_type='application/ms-excel')
                    response['Content-Disposition'] = 'attachment; filename="kardexproducto.xls"'
                    font_style = xlwt.XFStyle()
                    font_style.font.bold = True
                    fuentecabecera = easyxf('font: name Calibri, color-index black, bold on; pattern: pattern solid, fore_colour gray25; alignment: horiz centre; borders: left thin, right thin, top thin, bottom thin')
                    style2 = easyxf('borders: left thin, right thin, top thin, bottom thin; alignment: horiz centre')

                    # Estilos con un azul más oscuro y opaco
                    style_header_1 = easyxf('font: name Calibri, bold on, height 240, color-index white; alignment: horiz centre, vert centre; pattern: pattern solid, fore_colour 23; borders: left thin, right thin, top thin')
                    style_header_2 = easyxf('font: name Calibri, height 200, color-index white; alignment: horiz centre, vert centre; pattern: pattern solid, fore_colour 23; borders: left thin, right thin, bottom thin')

                    wb = xlwt.Workbook(encoding='utf-8')
                    ws = wb.add_sheet('hoja1')
                    row_num = 5
                    columns = [
                        ('ID', 2000),
                        ('Stock', 12000),
                        ('Movimiento', 8000),
                        ('Cantidad', 4000),
                        ('Costo', 4000),
                        ('Total', 4000),
                        ('Observación', 8000),
                    ]

                    # Escribir el título en las filas combinadas
                    ws.write_merge(0, 1, 0, 6, 'CLÍNICA SANTA ELENA', style_header_1)
                    ws.write_merge(2, 3, 0, 6, 'FLUJO DE MOVIMIENTOS DEL PRODUCTO EN STOCK', style_header_2)
                    for producto_ in Producto.objects.filter(status=True):
                        for stock_ in Stock.objects.filter(status=True, producto=producto_):
                            listado = KardexProducto.objects.filter(status=True, stock=stock_, fecha_creacion__gte=lunes_pasado, fecha_creacion__lte=ultimo_domingo).order_by('-id')
                            for col_num in range(len(columns)):
                                ws.write(row_num, col_num, columns[col_num][0], fuentecabecera)
                                ws.col(col_num).width = columns[col_num][1]

                            row_num += 1
                            for lis in listado.order_by('-fecha_creacion'):
                                ws.write(row_num, 0, str(lis.id), style2)
                                ws.write(row_num, 1, str(lis.stock), style2)
                                ws.write(row_num, 2, lis.get_movimiento_display(), style2)
                                ws.write(row_num, 3, lis.cantidad, style2)
                                ws.write(row_num, 4, f"{lis.costo:.2f}", style2)
                                ws.write(row_num, 5, f"{lis.total:.2f}", style2)
                                ws.write(row_num, 6, str(lis.observacion or ' '), style2)
                                row_num += 1
                            row_num += 1
                    wb.save(response)
                    return response
                except Exception as ex:
                    return JsonResponse({'result': False, 'error': str(ex)})

        else:
            try:
                data['title'] = u'Productos'
                request.session['viewactivo'] = 1
                # data['listado'] = Producto.objects.filter(status=True).order_by('-fecha_modificacion')
                search, url_vars, filtro = request.GET.get('s', ''), '', Q(status=True)
                if search:
                    data['s'] = search
                    filtro = filtro & \
                             Q(descripcion__icontains=search)\
                             |Q(nombre__icontains=search)\
                             |Q(subcategoria__descripcion__icontains=search)\
                             |Q(subcategoria__categoriaproducto__descripcion__icontains=search)
                    url_vars += f'&s={search}'
                listado = Producto.objects.filter(filtro).order_by('-id')
                for producto in listado:
                    producto.stock_total = producto.stocks.filter(status=True).aggregate(Sum('cantidad'))['cantidad__sum'] or 0
                paging = MiPaginador(listado, 10)
                p = 1
                try:
                    paginasesion = 1
                    if 'paginador' in request.session:
                        paginasesion = int(request.session['paginador'])
                    if 'page' in request.GET:
                        p = int(request.GET['page'])
                        # p = int(request.GET['page'])
                    else:
                        p = paginasesion
                    try:
                        page = paging.page(p)
                    except:
                        p = 1
                    page = paging.page(p)
                except:
                    page = paging.page(p)
                request.session['paginador'] = p
                data['paging'] = paging
                data["url_vars"] = url_vars
                data['rangospaging'] = paging.rangos_paginado(p)
                data['page'] = page
                data['listado'] = page.object_list
                return render(request, 'inventario/adm_productotiendavirtual/view.html', data)
            except Exception as ex:
                pass


def normalizar_texto(texto):
    return unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('ASCII').upper()


def comprobar_duplicado(producto, nuevas_especificaciones):
    """
    Verifica si ya existe un stock con el mismo producto y las mismas especificaciones.
    """
    stocks_existentes = Stock.objects.filter(producto=producto, status=True)
    nuevas_especificaciones = set(especificacion.id for especificacion in nuevas_especificaciones)

    for stock_existente in stocks_existentes:
        especificaciones_existentes = set(
            stock_existente.especificacion_producto.values_list('id', flat=True)
        )
        if nuevas_especificaciones == especificaciones_existentes:
            return True  # Duplicado encontrado
    return False  # No hay duplicados


def validar_especificacion(producto, especificaciones):
    """
    Valida que se hayan agregado todos los tipos de especificaciones requeridos para un producto.
    """
    count = []
    esp = producto.especificaciones.filter(status=True)

    if esp.exists():
        for especificacion in esp:
            tipo_especificacion = TipoEspecificacion.objects.get(pk=especificacion.especificacion_id)
            count.append(tipo_especificacion)

    tipos_especificaciones_requeridos = len(set(count))

    if tipos_especificaciones_requeridos != len(especificaciones):
        return True

    return False
